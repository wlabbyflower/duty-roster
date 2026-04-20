from __future__ import annotations

import json
import re
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from threading import RLock
from typing import Any
from uuid import uuid4
from zoneinfo import ZoneInfo

import requests
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
STATIC_DIR = BASE_DIR / "static"
SCHEDULE_FILE = DATA_DIR / "schedule.json"
SETTINGS_FILE = DATA_DIR / "settings.json"

DEFAULT_SETTINGS: dict[str, Any] = {
    "webhook_url": "",
    "notify_time": "09:00",
    "notify_count": 0,
    "mention_userids": [],
    "timezone": "Asia/Shanghai",
}

DATE_HEADERS = {"日期", "值班日期", "date", "day", "值班日"}
TIME_HEADERS = {"时间", "时段", "班次", "time", "shift"}
IGNORE_PERSON_HEADERS = {"", "人员"}
TIME_RE = re.compile(r"^([01]\d|2[0-3]):([0-5]\d)$")
SHEET_MONTH_RE = re.compile(r"^(\d{1,2})月$")
DAY_ONLY_RE = re.compile(r"(\d{1,2})\s*(日|号)?$")
MONTH_DAY_RE = re.compile(r"(\d{1,2})\s*[./月-]\s*(\d{1,2})")
FULL_DATE_RANGE_RE = re.compile(
    r"(?:(\d{4})\s*[年./-])?\s*(\d{1,2})\s*[月./-]\s*(\d{1,2})\s*(?:日|号)?\s*(?:到|至|[-~～—–])\s*"
    r"(?:(\d{4})\s*[年./-])?\s*(\d{1,2})\s*[月./-]\s*(\d{1,2})\s*(?:日|号)?"
)
SAME_MONTH_DATE_RANGE_RE = re.compile(
    r"(?:(\d{4})\s*[年./-])?\s*(\d{1,2})\s*[月./-]\s*(\d{1,2})\s*(?:日|号)?\s*(?:到|至|[-~～—–])\s*(\d{1,2})\s*(?:日|号)?"
)

HOLIDAY_NAME_KEYWORDS = ("节", "假", "春节", "清明", "劳动", "端午", "中秋", "国庆", "元旦")

WEEKDAY_MAP = {
    "周一": 1,
    "星期一": 1,
    "周二": 2,
    "星期二": 2,
    "周三": 3,
    "星期三": 3,
    "周四": 4,
    "星期四": 4,
    "周五": 5,
    "星期五": 5,
    "周六": 6,
    "星期六": 6,
    "周日": 7,
    "周天": 7,
    "星期日": 7,
    "星期天": 7,
}

app = FastAPI(title="Duty Roster")
lock = RLock()
scheduler = BackgroundScheduler()


class ScheduleRow(BaseModel):
    date: str
    pre_sales: str = Field(default="")
    after_sales: str = Field(default="")


class SaveSchedulePayload(BaseModel):
    rows: list[ScheduleRow]


class SaveSettingsPayload(BaseModel):
    webhook_url: str = ""
    notify_time: str = "09:00"
    notify_count: int = 0
    mention_userids: list[str] = Field(default_factory=list)
    timezone: str = "Asia/Shanghai"


class HolidayPeriodRow(BaseModel):
    id: str = ""
    name: str = ""
    start_date: str
    end_date: str
    pre_sales: str = Field(default="")
    after_sales: str = Field(default="")


DATA_DIR.mkdir(parents=True, exist_ok=True)


def _atomic_write(path: Path, value: Any) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(path)


def default_schedule_data() -> dict[str, Any]:
    return {
        "rows": [],
        "weekly_templates": [],
        "holiday_periods": [],
    }


def normalize_header(value: Any) -> str:
    return str(value or "").strip().replace(" ", "").replace("_", "").replace("-", "").lower()


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def append_unique(items: list[str], value: str) -> None:
    v = value.strip()
    if v and v not in items:
        items.append(v)


def is_pre_sales_assignment(role_text: str) -> bool:
    return "售前" in role_text or "推特私信" in role_text


def is_after_sales_assignment(role_text: str) -> bool:
    return "售后" in role_text


def normalize_detail_items(raw_items: Any) -> list[dict[str, str]]:
    if not isinstance(raw_items, list):
        return []
    result: list[dict[str, str]] = []
    for item in raw_items:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name") or "").strip()
        status = str(item.get("status") or "").strip()
        if not name:
            continue
        if not status:
            status = "值班中"
        result.append({"name": name, "status": status})
    return result


def normalize_iso_date_string(value: Any) -> str | None:
    d = normalize_date_value(value)
    if not d:
        return None
    return d.isoformat()


def normalize_date_value(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    s = str(value).strip()
    if not s:
        return None

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def validate_time(value: str) -> str:
    if not TIME_RE.match(value or ""):
        raise HTTPException(status_code=400, detail="通知时间格式应为 HH:MM")
    return value


def validate_timezone(value: str) -> str:
    try:
        ZoneInfo(value)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail="无效时区") from exc
    return value


def load_settings() -> dict[str, Any]:
    with lock:
        if not SETTINGS_FILE.exists():
            _atomic_write(SETTINGS_FILE, DEFAULT_SETTINGS)
            return dict(DEFAULT_SETTINGS)

        raw = json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
        settings = dict(DEFAULT_SETTINGS)
        settings.update(raw if isinstance(raw, dict) else {})

        settings["notify_time"] = settings.get("notify_time") or DEFAULT_SETTINGS["notify_time"]
        settings["notify_count"] = max(0, int(settings.get("notify_count") or 0))
        settings["mention_userids"] = [str(i).strip() for i in settings.get("mention_userids") or [] if str(i).strip()]
        settings["timezone"] = settings.get("timezone") or DEFAULT_SETTINGS["timezone"]
        return settings


def save_settings(settings: dict[str, Any]) -> None:
    with lock:
        _atomic_write(SETTINGS_FILE, settings)


def normalize_schedule_rows(raw_rows: list[Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in raw_rows:
        if not isinstance(item, dict):
            continue
        d = normalize_date_value(item.get("date"))
        if not d:
            continue
        pre_sales = str(item.get("pre_sales") or "").strip()
        after_sales = str(item.get("after_sales") or "").strip()
        pre_details = normalize_detail_items(item.get("pre_details"))
        after_details = normalize_detail_items(item.get("after_details"))
        rows.append(
            {
                "date": d.isoformat(),
                "pre_sales": pre_sales,
                "after_sales": after_sales,
                "pre_details": pre_details,
                "after_details": after_details,
            }
        )

    by_date: dict[str, dict[str, str]] = {}
    for row in rows:
        by_date[row["date"]] = row
    result = list(by_date.values())
    result.sort(key=lambda x: x["date"])
    return result


def normalize_weekly_templates(raw_templates: list[Any]) -> list[dict[str, Any]]:
    templates: list[dict[str, Any]] = []
    for item in raw_templates:
        if not isinstance(item, dict):
            continue
        weekday = int(item.get("weekday") or 0)
        if weekday < 1 or weekday > 7:
            continue
        pre_sales = str(item.get("pre_sales") or "").strip()
        after_sales = str(item.get("after_sales") or "").strip()
        pre_details = normalize_detail_items(item.get("pre_details"))
        after_details = normalize_detail_items(item.get("after_details"))
        templates.append(
            {
                "weekday": weekday,
                "pre_sales": pre_sales,
                "after_sales": after_sales,
                "pre_details": pre_details,
                "after_details": after_details,
            }
        )

    by_weekday: dict[int, dict[str, Any]] = {}
    for tpl in templates:
        by_weekday[tpl["weekday"]] = tpl
    result = list(by_weekday.values())
    result.sort(key=lambda x: x["weekday"])
    return result


def normalize_holiday_periods(raw_periods: list[Any]) -> list[dict[str, Any]]:
    periods: list[dict[str, Any]] = []
    for item in raw_periods:
        if not isinstance(item, dict):
            continue

        start_date = normalize_iso_date_string(item.get("start_date"))
        end_date = normalize_iso_date_string(item.get("end_date"))
        if not start_date or not end_date:
            continue
        if start_date > end_date:
            start_date, end_date = end_date, start_date

        periods.append(
            {
                "id": str(item.get("id") or uuid4()),
                "name": str(item.get("name") or "").strip(),
                "start_date": start_date,
                "end_date": end_date,
                "range_start_date": normalize_iso_date_string(item.get("range_start_date")) or start_date,
                "range_end_date": normalize_iso_date_string(item.get("range_end_date")) or end_date,
                "pre_sales": str(item.get("pre_sales") or "").strip(),
                "after_sales": str(item.get("after_sales") or "").strip(),
                "pre_details": normalize_detail_items(item.get("pre_details")),
                "after_details": normalize_detail_items(item.get("after_details")),
            }
        )

    unique: dict[tuple[str, str, str, str, str], dict[str, Any]] = {}
    for period in periods:
        key = (
            period["start_date"],
            period["end_date"],
            period["name"],
            period["range_start_date"],
            period["range_end_date"],
        )
        unique[key] = period

    result = list(unique.values())
    result.sort(key=lambda x: (x["start_date"], x["end_date"], x["name"]))
    return result


def load_schedule_data() -> dict[str, Any]:
    with lock:
        if not SCHEDULE_FILE.exists():
            data = default_schedule_data()
            _atomic_write(SCHEDULE_FILE, data)
            return data

        raw = json.loads(SCHEDULE_FILE.read_text(encoding="utf-8"))

        if isinstance(raw, list):
            rows = normalize_schedule_rows(raw)
            data = {"rows": rows, "weekly_templates": [], "holiday_periods": []}
            _atomic_write(SCHEDULE_FILE, data)
            return data

        if not isinstance(raw, dict):
            data = default_schedule_data()
            _atomic_write(SCHEDULE_FILE, data)
            return data

        rows = normalize_schedule_rows(raw.get("rows") or [])
        weekly_templates = normalize_weekly_templates(raw.get("weekly_templates") or [])
        holiday_periods = normalize_holiday_periods(raw.get("holiday_periods") or [])
        data = {"rows": rows, "weekly_templates": weekly_templates, "holiday_periods": holiday_periods}
        return data


def save_schedule_data(data: dict[str, Any]) -> None:
    with lock:
        _atomic_write(SCHEDULE_FILE, data)


def load_schedule_rows() -> list[dict[str, str]]:
    return load_schedule_data()["rows"]


def save_schedule_rows(rows: list[dict[str, str]]) -> None:
    data = load_schedule_data()
    data["rows"] = normalize_schedule_rows(rows)
    save_schedule_data(data)


def save_holiday_periods(periods: list[dict[str, Any]]) -> None:
    data = load_schedule_data()
    data["holiday_periods"] = normalize_holiday_periods(periods)
    save_schedule_data(data)


def parse_schedule_rows(rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    normalized: list[dict[str, Any]] = []
    for row in rows:
        d = normalize_date_value(row.get("date"))
        if not d:
            continue
        pre_sales = str(row.get("pre_sales") or "").strip()
        after_sales = str(row.get("after_sales") or "").strip()
        if not pre_sales and not after_sales:
            continue
        normalized.append({"date": d.isoformat(), "pre_sales": pre_sales, "after_sales": after_sales})

    by_date: dict[str, dict[str, str]] = {r["date"]: r for r in normalized}
    result = list(by_date.values())
    result.sort(key=lambda x: x["date"])
    return result


def parse_holiday_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for row in rows:
        start_date = normalize_iso_date_string(row.get("start_date"))
        end_date = normalize_iso_date_string(row.get("end_date"))
        if not start_date or not end_date:
            continue
        if start_date > end_date:
            start_date, end_date = end_date, start_date

        pre_sales = str(row.get("pre_sales") or "").strip()
        after_sales = str(row.get("after_sales") or "").strip()
        if not pre_sales and not after_sales:
            continue

        normalized.append(
            {
                "id": str(row.get("id") or uuid4()),
                "name": str(row.get("name") or "").strip(),
                "start_date": start_date,
                "end_date": end_date,
                "range_start_date": normalize_iso_date_string(row.get("range_start_date")) or start_date,
                "range_end_date": normalize_iso_date_string(row.get("range_end_date")) or end_date,
                "pre_sales": pre_sales,
                "after_sales": after_sales,
            }
        )

    unique: dict[tuple[str, str, str, str, str], dict[str, Any]] = {}
    for row in normalized:
        unique[(row["start_date"], row["end_date"], row["name"], row["range_start_date"], row["range_end_date"])] = row
    result = list(unique.values())
    result.sort(key=lambda x: (x["start_date"], x["end_date"], x["name"]))
    return result


def parse_sheet_month(sheet_name: str) -> int | None:
    m = SHEET_MONTH_RE.match(sheet_name.strip())
    if not m:
        return None
    month = int(m.group(1))
    if 1 <= month <= 12:
        return month
    return None


def infer_year_for_month(month: int, today: date) -> int:
    candidates = [today.year - 1, today.year, today.year + 1]
    best_year = today.year
    best_diff = None
    for y in candidates:
        try:
            d = date(y, month, 15)
        except ValueError:
            continue
        diff = abs((d - today).days)
        if best_diff is None or diff < best_diff:
            best_diff = diff
            best_year = y
    return best_year


def parse_weekday(value: Any) -> int | None:
    s = str(value or "").strip().replace(" ", "")
    if not s:
        return None
    return WEEKDAY_MAP.get(s)


def parse_day_descriptor(value: Any, sheet_month: int | None, today: date) -> tuple[str, Any] | None:
    weekday = parse_weekday(value)
    if weekday:
        return ("weekday", weekday)

    d = normalize_date_value(value)
    if d:
        return ("date", d.isoformat())

    s = str(value or "").strip()
    if not s:
        return None

    m = MONTH_DAY_RE.search(s)
    if m:
        month = int(m.group(1))
        day_num = int(m.group(2))
        if 1 <= month <= 12 and 1 <= day_num <= 31:
            year = infer_year_for_month(month, today)
            try:
                d = date(year, month, day_num)
                return ("date", d.isoformat())
            except ValueError:
                pass

    m2 = DAY_ONLY_RE.search(s)
    if m2 and sheet_month:
        day_num = int(m2.group(1))
        if 1 <= day_num <= 31:
            year = infer_year_for_month(sheet_month, today)
            try:
                d = date(year, sheet_month, day_num)
                return ("date", d.isoformat())
            except ValueError:
                pass

    return None


def parse_time_slot(value: Any) -> str:
    s = str(value or "").strip().replace("：", ":")
    if not s:
        return "unknown"
    m = re.search(r"(\d{1,2})\s*:\s*\d{1,2}", s)
    if not m:
        return "unknown"
    hour = int(m.group(1))
    if hour < 12:
        return "morning"
    if hour < 18:
        return "afternoon"
    return "evening"


def excel_serial_to_date(serial: float) -> date | None:
    try:
        epoch = datetime(1899, 12, 30)
        return (epoch + timedelta(days=float(serial))).date()
    except Exception:  # noqa: BLE001
        return None


def sort_people_by_order(candidates: list[str], person_order: list[str]) -> list[str]:
    seen = set()
    ordered: list[str] = []
    for name in person_order:
        if name in candidates and name not in seen:
            ordered.append(name)
            seen.add(name)
    for name in candidates:
        if name not in seen:
            ordered.append(name)
            seen.add(name)
    return ordered


def collect_sheet_strings(rows: list[list[Any]], limit: int = 8) -> str:
    parts: list[str] = []
    for row in rows[:limit]:
        for cell in row:
            text = normalize_text(cell)
            if text:
                parts.append(text)
    return " ".join(parts)


def infer_year_for_holiday_range(start_month: int, end_month: int, today: date) -> int:
    candidate_month = start_month if start_month >= today.month else end_month
    return infer_year_for_month(candidate_month, today)


def parse_holiday_range_text(text: str, today: date) -> tuple[str, str] | None:
    s = normalize_text(text).replace("～", "~").replace("—", "-").replace("–", "-")
    if not s:
        return None

    match = FULL_DATE_RANGE_RE.search(s)
    if match:
        start_year = int(match.group(1)) if match.group(1) else None
        start_month = int(match.group(2))
        start_day = int(match.group(3))
        end_year = int(match.group(4)) if match.group(4) else None
        end_month = int(match.group(5))
        end_day = int(match.group(6))
        year = start_year or end_year or infer_year_for_holiday_range(start_month, end_month, today)
        try:
            start = date(start_year or year, start_month, start_day)
            end = date(end_year or year, end_month, end_day)
            return (start.isoformat(), end.isoformat()) if start <= end else (end.isoformat(), start.isoformat())
        except ValueError:
            return None

    match = SAME_MONTH_DATE_RANGE_RE.search(s)
    if match:
        explicit_year = int(match.group(1)) if match.group(1) else None
        month = int(match.group(2))
        start_day = int(match.group(3))
        end_day = int(match.group(4))
        year = explicit_year or infer_year_for_month(month, today)
        try:
            start = date(year, month, start_day)
            end = date(year, month, end_day)
            return (start.isoformat(), end.isoformat()) if start <= end else (end.isoformat(), start.isoformat())
        except ValueError:
            return None

    return None


def detect_holiday_title(rows: list[list[Any]]) -> str:
    for row in rows[:6]:
        for cell in row:
            text = normalize_text(cell)
            if not text:
                continue
            if any(keyword in text for keyword in HOLIDAY_NAME_KEYWORDS):
                return text
    return ""


def detect_holiday_date_range(rows: list[list[Any]], today: date) -> tuple[str, str] | None:
    for row in rows[:8]:
        for cell in row:
            text = normalize_text(cell)
            if not text:
                continue
            parsed = parse_holiday_range_text(text, today)
            if parsed:
                return parsed
    return None


def parse_holiday_name_from_range_text(rows: list[list[Any]], fallback: str) -> str:
    title = detect_holiday_title(rows)
    if title:
        return title
    return fallback or "节假日值班"


def date_range_iter(start_iso: str, end_iso: str):
    start = datetime.strptime(start_iso, "%Y-%m-%d").date()
    end = datetime.strptime(end_iso, "%Y-%m-%d").date()
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def save_single_sheet_to_bytes(ws) -> bytes:
    wb = Workbook()
    target = wb.active
    target.title = ws.title
    for row in ws.iter_rows(values_only=True):
        target.append(list(row))
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def summarize_slot_status(slots: set[str]) -> str:
    s = set(slots)
    if not s:
        return "值班中"
    if "unknown" in s:
        s.discard("unknown")
        if not s:
            return "全天都在"

    if s in ({"morning"},):
        return "上午在"
    if s in ({"afternoon"},):
        return "下午在"
    if s in ({"evening"},):
        return "晚上在"
    if s == {"morning", "afternoon"}:
        return "白天在"
    if s == {"afternoon", "evening"}:
        return "下午来晚上也在"
    if s == {"morning", "afternoon", "evening"}:
        return "全天都在"
    if s == {"morning", "evening"}:
        return "上午和晚上在"
    return "值班中"


def build_role_output(
    bucket: dict[str, Any],
    role: str,
    role_roster: set[str],
    person_order: list[str],
) -> tuple[str, list[dict[str, str]]]:
    work_map: dict[str, set[str]] = bucket.get(f"{role}_work", {})
    rest_set: set[str] = bucket.get("rest", set())

    work_people = sort_people_by_order(list(work_map.keys()), person_order)
    details: list[dict[str, str]] = [
        {"name": person, "status": summarize_slot_status(work_map.get(person, set()))}
        for person in work_people
    ]

    rest_people = sort_people_by_order(
        [p for p in rest_set if p in role_roster and p not in work_map],
        person_order,
    )
    for person in rest_people:
        details.append({"name": person, "status": "今日休息"})

    return "、".join(work_people), details


def detect_header_row(rows: list[list[Any]]) -> tuple[int, int, int, list[tuple[int, str]]] | None:
    scan_limit = min(len(rows), 30)

    for i in range(scan_limit):
        row = rows[i]
        normalized = [normalize_header(c) for c in row]

        date_col = None
        time_col = None
        for idx, h in enumerate(normalized):
            if date_col is None and h in DATE_HEADERS:
                date_col = idx
            if time_col is None and h in TIME_HEADERS:
                time_col = idx

        if date_col is None or time_col is None:
            continue

        person_cols: list[tuple[int, str]] = []
        seen_person = False
        blank_after_person = 0
        for idx in range(time_col + 1, len(row)):
            raw_val = row[idx]
            name = normalize_text(raw_val)

            if not name:
                if seen_person:
                    blank_after_person += 1
                    if blank_after_person >= 1:
                        break
                continue

            blank_after_person = 0

            normalized_name = normalize_header(name)
            if normalized_name in IGNORE_PERSON_HEADERS:
                continue

            # Filter out numeric/date-like headers (e.g. 10.2 / 10-2) from side-by-side tables.
            if isinstance(raw_val, (int, float)):
                continue
            if re.fullmatch(r"\d{1,2}([./-]\d{1,2})?", name):
                continue

            seen_person = True
            person_cols.append((idx, name))

        if person_cols:
            return (i, date_col, time_col, person_cols)

    return None


def parse_excel_schedule(content: bytes, tz_name: str) -> dict[str, Any]:
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    today = datetime.now(ZoneInfo(tz_name)).date()

    date_map: dict[str, dict[str, Any]] = {}
    weekday_map: dict[int, dict[str, Any]] = {}
    person_order: list[str] = []
    pre_roster: set[str] = set()
    after_roster: set[str] = set()

    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            continue

        header = detect_header_row(rows)
        if not header:
            continue

        header_idx, date_col, time_col, person_cols = header
        sheet_month = parse_sheet_month(ws.title)
        current_day_marker: Any = None
        for _, person_name in person_cols:
            append_unique(person_order, person_name)

        for row in rows[header_idx + 1 :]:
            if date_col < len(row) and normalize_text(row[date_col]):
                current_day_marker = row[date_col]

            if current_day_marker is None:
                continue

            descriptor = parse_day_descriptor(current_day_marker, sheet_month, today)
            if not descriptor:
                continue

            has_assignment = False
            slot = parse_time_slot(row[time_col] if time_col < len(row) else None)

            for idx, person_name in person_cols:
                if idx >= len(row):
                    continue
                role_text = normalize_text(row[idx])
                if not role_text:
                    continue
                has_assignment = True

            if not has_assignment:
                continue

            kind, key = descriptor
            if kind == "date":
                bucket = date_map.setdefault(
                    str(key),
                    {
                        "pre_work": {},
                        "after_work": {},
                        "rest": set(),
                    },
                )
            else:
                weekday = int(key)
                bucket = weekday_map.setdefault(
                    weekday,
                    {
                        "pre_work": {},
                        "after_work": {},
                        "rest": set(),
                    },
                )

            for idx, person_name in person_cols:
                if idx >= len(row):
                    continue
                role_text = normalize_text(row[idx])
                if not role_text:
                    continue

                if "休息" in role_text:
                    bucket["rest"].add(person_name)

                if is_pre_sales_assignment(role_text):
                    pre_roster.add(person_name)
                    work_map = bucket["pre_work"]
                    work_map.setdefault(person_name, set()).add(slot)

                if is_after_sales_assignment(role_text):
                    after_roster.add(person_name)
                    work_map = bucket["after_work"]
                    work_map.setdefault(person_name, set()).add(slot)

    rows_result = []
    for d, values in date_map.items():
        pre_sales, pre_details = build_role_output(values, "pre", pre_roster, person_order)
        after_sales, after_details = build_role_output(values, "after", after_roster, person_order)
        rows_result.append(
            {
                "date": d,
                "pre_sales": pre_sales,
                "after_sales": after_sales,
                "pre_details": pre_details,
                "after_details": after_details,
            }
        )
    rows_result.sort(key=lambda x: x["date"])

    weekly_templates = []
    for weekday, values in weekday_map.items():
        pre_sales, pre_details = build_role_output(values, "pre", pre_roster, person_order)
        after_sales, after_details = build_role_output(values, "after", after_roster, person_order)
        weekly_templates.append(
            {
                "weekday": weekday,
                "pre_sales": pre_sales,
                "after_sales": after_sales,
                "pre_details": pre_details,
                "after_details": after_details,
            }
        )
    weekly_templates.sort(key=lambda x: x["weekday"])

    if not rows_result and not weekly_templates:
        raise HTTPException(status_code=400, detail="未识别到可用值班数据，请检查 Excel 模板")

    return {
        "rows": rows_result,
        "weekly_templates": weekly_templates,
    }


def parse_holiday_excel_schedule(content: bytes, tz_name: str) -> dict[str, Any]:
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    today = datetime.now(ZoneInfo(tz_name)).date()

    periods: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            continue

        date_range = detect_holiday_date_range(rows, today)
        if not date_range:
            continue

        parsed = parse_excel_schedule(content=save_single_sheet_to_bytes(ws), tz_name=tz_name)
        period_name = parse_holiday_name_from_range_text(rows, ws.title.strip())

        if parsed.get("rows"):
            exact_rows = parsed["rows"]
            range_start_date, range_end_date = date_range
            for row in exact_rows:
                periods.append(
                    {
                        "name": period_name,
                        "start_date": row["date"],
                        "end_date": row["date"],
                        "range_start_date": range_start_date,
                        "range_end_date": range_end_date,
                        "pre_sales": row.get("pre_sales", ""),
                        "after_sales": row.get("after_sales", ""),
                        "pre_details": normalize_detail_items(row.get("pre_details")),
                        "after_details": normalize_detail_items(row.get("after_details")),
                    }
                )
        else:
            start_date, end_date = date_range
            templates = parsed.get("weekly_templates") or []
            if not templates:
                raise HTTPException(status_code=400, detail="节假日 Excel 已识别到日期区间，但未识别到排班内容")
            for current in date_range_iter(start_date, end_date):
                weekday = current.isoweekday()
                matched = next((tpl for tpl in templates if int(tpl.get("weekday") or 0) == weekday), None)
                if not matched:
                    continue
                periods.append(
                    {
                        "name": period_name,
                        "start_date": current.isoformat(),
                        "end_date": current.isoformat(),
                        "range_start_date": start_date,
                        "range_end_date": end_date,
                        "pre_sales": matched.get("pre_sales", ""),
                        "after_sales": matched.get("after_sales", ""),
                        "pre_details": normalize_detail_items(matched.get("pre_details")),
                        "after_details": normalize_detail_items(matched.get("after_details")),
                    }
                )

    periods = normalize_holiday_periods(periods)
    if not periods:
        raise HTTPException(status_code=400, detail="未识别到节假日值班数据，请确认表格顶部包含起止日期")
    return {"holiday_periods": periods}


def get_today_entry(schedule_data: dict[str, Any], tz_name: str) -> dict[str, Any] | None:
    now = datetime.now(ZoneInfo(tz_name))
    today_iso = now.date().isoformat()

    holiday_matches = [
        period
        for period in schedule_data.get("holiday_periods") or []
        if str(period.get("start_date") or "") <= today_iso <= str(period.get("end_date") or "")
    ]
    holiday_matches.sort(
        key=lambda item: (
            item.get("start_date") != item.get("end_date"),
            item.get("start_date", ""),
            item.get("end_date", ""),
        )
    )
    if holiday_matches:
        period = holiday_matches[0]
        return {
            "date": today_iso,
            "pre_sales": period.get("pre_sales", ""),
            "after_sales": period.get("after_sales", ""),
            "pre_details": normalize_detail_items(period.get("pre_details")),
            "after_details": normalize_detail_items(period.get("after_details")),
            "source": "holiday",
            "holiday_name": period.get("name", ""),
            "holiday_start_date": period.get("range_start_date", period.get("start_date", "")),
            "holiday_end_date": period.get("range_end_date", period.get("end_date", "")),
        }

    for row in schedule_data.get("rows") or []:
        if row.get("date") == today_iso:
            return {
                "date": row["date"],
                "pre_sales": row.get("pre_sales", ""),
                "after_sales": row.get("after_sales", ""),
                "pre_details": normalize_detail_items(row.get("pre_details")),
                "after_details": normalize_detail_items(row.get("after_details")),
                "source": "date",
            }

    weekday = now.isoweekday()
    for tpl in schedule_data.get("weekly_templates") or []:
        if int(tpl.get("weekday") or 0) == weekday:
            return {
                "date": today_iso,
                "pre_sales": tpl.get("pre_sales", ""),
                "after_sales": tpl.get("after_sales", ""),
                "pre_details": normalize_detail_items(tpl.get("pre_details")),
                "after_details": normalize_detail_items(tpl.get("after_details")),
                "source": "weekday",
            }

    return None


def send_wecom_notification(test_mode: bool = False) -> dict[str, Any]:
    settings = load_settings()
    webhook_url = settings.get("webhook_url", "").strip()
    if not webhook_url:
        return {"sent": False, "reason": "未配置企业微信机器人 Webhook"}

    schedule_data = load_schedule_data()
    tz_name = settings.get("timezone") or DEFAULT_SETTINGS["timezone"]
    today_entry = get_today_entry(schedule_data, tz_name)

    def role_lines(role_name: str, names: str, details: list[dict[str, str]]) -> list[str]:
        if details:
            lines = [f"{role_name}："]
            for d in details:
                lines.append(f"{d['name']}，{d['status']}")
            return lines
        return [f"{role_name}：{names or '未安排'}"]

    if today_entry:
        pre_details = normalize_detail_items(today_entry.get("pre_details"))
        after_details = normalize_detail_items(today_entry.get("after_details"))

        is_holiday = today_entry.get("source") == "holiday"
        title = "节假日值班提醒" if is_holiday else "值班提醒"
        lines = [title, f"日期：{today_entry['date']}"]
        if is_holiday:
            holiday_name = str(today_entry.get("holiday_name") or "").strip()
            holiday_start = str(today_entry.get("holiday_start_date") or "")
            holiday_end = str(today_entry.get("holiday_end_date") or "")
            if holiday_name:
                lines.append(f"节假日：{holiday_name}")
            if holiday_start and holiday_end:
                lines.append(f"区间：{holiday_start} 至 {holiday_end}")
        lines.extend(role_lines("售前", today_entry.get("pre_sales", ""), pre_details))
        lines.extend(role_lines("售后", today_entry.get("after_sales", ""), after_details))
        content = "\n".join(lines)
    else:
        content = "值班提醒\n今日暂无排班"

    mention_ids = [str(i).strip() for i in settings.get("mention_userids") or [] if str(i).strip()]
    notify_count = max(0, int(settings.get("notify_count") or 0))
    mention_ids = mention_ids[:notify_count] if notify_count > 0 else []

    payload: dict[str, Any] = {
        "msgtype": "text",
        "text": {
            "content": content,
            "mentioned_list": mention_ids,
        },
    }

    try:
        resp = requests.post(webhook_url, json=payload, timeout=12)
        body = resp.json() if resp.headers.get("content-type", "").startswith("application/json") else {}
    except Exception as exc:  # noqa: BLE001
        return {"sent": False, "reason": str(exc)}

    success = resp.status_code == 200 and body.get("errcode") == 0
    return {
        "sent": success,
        "status_code": resp.status_code,
        "response": body,
        "test_mode": test_mode,
    }


def schedule_daily_job() -> None:
    settings = load_settings()
    notify_time = validate_time(settings.get("notify_time") or DEFAULT_SETTINGS["notify_time"])
    timezone = validate_timezone(settings.get("timezone") or DEFAULT_SETTINGS["timezone"])

    hour, minute = [int(x) for x in notify_time.split(":")]
    trigger = CronTrigger(hour=hour, minute=minute, timezone=timezone)

    if scheduler.get_job("daily_notify"):
        scheduler.remove_job("daily_notify")

    scheduler.add_job(lambda: send_wecom_notification(test_mode=False), trigger, id="daily_notify")


@app.on_event("startup")
def _startup() -> None:
    load_settings()
    load_schedule_data()
    if not scheduler.running:
        scheduler.start()
    schedule_daily_job()


@app.on_event("shutdown")
def _shutdown() -> None:
    if scheduler.running:
        scheduler.shutdown(wait=False)


@app.get("/")
def index() -> FileResponse:
    return FileResponse(STATIC_DIR / "index.html")


@app.get("/api/today")
def api_today() -> dict[str, Any]:
    settings = load_settings()
    tz_name = settings.get("timezone") or DEFAULT_SETTINGS["timezone"]
    schedule_data = load_schedule_data()
    entry = get_today_entry(schedule_data, tz_name)
    return {
        "today": entry,
        "timezone": tz_name,
    }


@app.get("/api/schedule")
def api_schedule() -> dict[str, Any]:
    data = load_schedule_data()
    return {
        "rows": data["rows"],
        "weekly_templates": data.get("weekly_templates") or [],
        "holiday_periods": data.get("holiday_periods") or [],
    }


@app.post("/api/schedule")
def api_save_schedule(payload: SaveSchedulePayload) -> dict[str, Any]:
    rows = parse_schedule_rows([r.model_dump() for r in payload.rows])
    save_schedule_rows(rows)
    return {"ok": True, "count": len(rows)}


@app.post("/api/import-excel")
async def api_import_excel(file: UploadFile = File(...)) -> dict[str, Any]:
    if not file.filename:
        raise HTTPException(status_code=400, detail="未选择文件")

    suffix = Path(file.filename).suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 格式")

    content = await file.read()
    settings = load_settings()
    tz_name = settings.get("timezone") or DEFAULT_SETTINGS["timezone"]
    parsed = parse_excel_schedule(content, tz_name)
    save_schedule_data(parsed)
    return {
        "ok": True,
        "count": len(parsed["rows"]),
        "weekly_count": len(parsed.get("weekly_templates") or []),
    }


@app.post("/api/import-holiday-excel")
async def api_import_holiday_excel(file: UploadFile = File(...)) -> dict[str, Any]:
    if not file.filename:
        raise HTTPException(status_code=400, detail="未选择文件")

    suffix = Path(file.filename).suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 格式")

    content = await file.read()
    settings = load_settings()
    tz_name = settings.get("timezone") or DEFAULT_SETTINGS["timezone"]
    parsed = parse_holiday_excel_schedule(content, tz_name)
    save_holiday_periods(parsed["holiday_periods"])
    return {
        "ok": True,
        "count": len(parsed["holiday_periods"]),
    }


@app.get("/api/export-excel")
def api_export_excel() -> StreamingResponse:
    data = load_schedule_data()
    rows = data["rows"]
    weekly = data.get("weekly_templates") or []
    holiday_periods = data.get("holiday_periods") or []

    wb = Workbook()
    ws = wb.active
    ws.title = "按日期值班"
    ws.append(["日期", "售前", "售后"])
    for r in rows:
        ws.append([r["date"], r["pre_sales"], r["after_sales"]])

    ws2 = wb.create_sheet("按周模板")
    ws2.append(["周几", "售前", "售后"])
    weekday_name = {1: "周一", 2: "周二", 3: "周三", 4: "周四", 5: "周五", 6: "周六", 7: "周日"}
    for tpl in weekly:
        w = int(tpl.get("weekday") or 0)
        ws2.append([weekday_name.get(w, w), tpl.get("pre_sales", ""), tpl.get("after_sales", "")])

    ws3 = wb.create_sheet("节假日值班")
    ws3.append(["名称", "开始日期", "结束日期", "售前", "售后"])
    for item in holiday_periods:
        ws3.append(
            [
                item.get("name", ""),
                item.get("start_date", ""),
                item.get("end_date", ""),
                item.get("pre_sales", ""),
                item.get("after_sales", ""),
            ]
        )

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=duty_roster.xlsx"},
    )


@app.get("/api/settings")
def api_get_settings() -> dict[str, Any]:
    return load_settings()


@app.post("/api/settings")
def api_save_settings(payload: SaveSettingsPayload) -> dict[str, Any]:
    data = payload.model_dump()
    data["notify_time"] = validate_time(data.get("notify_time") or DEFAULT_SETTINGS["notify_time"])
    data["timezone"] = validate_timezone(data.get("timezone") or DEFAULT_SETTINGS["timezone"])
    data["notify_count"] = max(0, int(data.get("notify_count") or 0))
    data["mention_userids"] = [str(i).strip() for i in data.get("mention_userids") or [] if str(i).strip()]

    save_settings(data)
    schedule_daily_job()
    return {"ok": True}


@app.post("/api/notify/test")
def api_notify_test() -> JSONResponse:
    result = send_wecom_notification(test_mode=True)
    status = 200 if result.get("sent") else 400
    return JSONResponse(result, status_code=status)


app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
